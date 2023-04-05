import gc
from tqdm import tqdm
import os
import pandas as pd

# def merge_csv_to_excel():
#     path = f'./최종모음'
#     # 현재 폴더 내의 csv 파일 리스트를 가져옴
#     csv_files = [f for f in os.listdir(path) if f.endswith('.csv')]
    
#     # csv 파일을 순회하며 데이터를 병합
#     writer = pd.ExcelWriter('merged.xlsx', engine='xlsxwriter')
#     for csv_file in tqdm(csv_files):
#         sheet_name = csv_file.split('.')[0]
#         df = pd.read_csv(f'{path}/{csv_file}', index_col=0, header = None ,delimiter=" " ,skip_blank_lines=True )
#         df.to_excel(writer, sheet_name=sheet_name, index=False)
    
#     del df
#     gc.collect()

#     # 작성한 엑셀 파일 저장
#     writer.save()


import openpyxl
import csv
import os

# 결과 파일 이름
result_file_name = 'merged.xlsx'

# csv 파일이 있는 폴더 경로
csv_folder_path = './최종모음'

# xlsx 파일 생성
result_wb = openpyxl.Workbook()

# csv 파일들을 읽어들여서 각각의 시트로 추가
for file_name in os.listdir(csv_folder_path):
    if file_name.endswith('.csv'):
        sheet_name = file_name.replace('.csv', '')
        with open(os.path.join(csv_folder_path, file_name), 'r') as f:
            csv_reader = csv.reader(f)
            sheet = result_wb.create_sheet(title=sheet_name)
            for r_idx, row in enumerate(csv_reader):
                for c_idx, value in enumerate(row):
                    sheet.cell(row=r_idx+1, column=c_idx+1).value = value

# 첫번째 시트 제거
result_wb.remove(result_wb['Sheet'])

# 결과 파일 저장
result_wb.save(result_file_name)


# if __name__ == "__main__" : 
#     merge_csv_to_excel()